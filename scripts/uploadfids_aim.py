import time

import openpyxl
import requests
import os
from requests import JSONDecodeError


def findExcel(file):
    """Method -> is a void
    1. loads a given file
    2. refer to sheet1 and column A
    3. prints each cell value
      """
    try:
        workbook = openpyxl.load_workbook(file)
        if workbook is not None:
            sheet = workbook['Sheet1']
            count = 1
            for cell in sheet['A']:
                print(f'Invoking API for FID - ', {cell.value})
                api_response = api_call(cell.value)
                if api_response is not None:
                    write_api_response_file(api_response, count)
                    count += 1
                time.sleep(5)

        print("end of program!!")

    except Exception as exception:
        print("exception occurred - ", exception)


def api_call(message):

    """function - api_call
       :parameter -> message
       returns response from API"""

    url = 'http://localhost:8080/api/post-message'

    api_request_json = {"message": message}
    headers = {"Accept": "application/json", "application-id": "105818"}

    try:
        print(f'API request for Jet ID', {message})
        response = requests.post(url, api_request_json, headers=headers)
        response_json = str(response.content)
        return response_json

    except JSONDecodeError as error:
        raise Exception('Json conversion failed', error)

    except Exception as exception:
        return exception


def write_api_response_file(api_response, count):
    """
    :param api_response: string
    :param count:
    :return: void

    creates an output file
    and writes the api_response to the file
    """

    os.chdir('C:\\Sample\\PycharmProjects\\automate-boaringstuff')
    output_file = 'out_put_file.xlsx'

    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.cell(row=count, column=1, value=str(api_response))
    wb.save(output_file)


if __name__ == '__main__':
    findExcel('jetfids.xlsx')
